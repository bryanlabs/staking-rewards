import requests
import time
import sys
import json
import argparse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from os.path import exists
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import traceback
import pydoc
import csv

COIN_GECKO_API_URL = "https://api.coingecko.com/api/v3"
COIN_GECKO_COINS_ENDPOINT = "/coins"
COIN_GECKO_HISTORICAL_ENDPOINT = "/<coin-id>/history"
COIN_GECKO_COINS_LIST_ENDPOINT = "/list"

#this represents CoinGecko's 50 requests/minute API rate limit, plus some seconds for variance
COIN_GECKO_REQUEST_CHUNKS = 9
COIN_GECKO_THROTTLE_TIME = 61

#this represents throttling for Coinhall, which will throttle after every single request
COINHALL_THROTTLE_TIME = 10
COINHALL_API_URL = "https://api.coinhall.org/api/v1"
COINHALL_CHART_ENDPOINT = "/charts/terra/candles"
COINHALL_HISTORICAL_PARAMS = "?bars=1&from=<from-time>&to=<to-time>&quoteAsset=uusd&interval=1d&pairAddress=<coin-id>"

KEY_CLASSIFICATIONS = ["staked", "airdrop"]

class CaughtError(Exception):
    pass

def parse_args(process_function, import_symbol_function):
    parser = argparse.ArgumentParser(description='A command line tool to process CSV/XLSX files of Crypto Symbols and amounts and gather USD value')
    subparser = parser.add_subparsers()

    parser_process_file = subparser.add_parser("process")
    parser_process_file.add_argument('--input-file', '-i', help='The location of your Accointing input file', required=True)
    parser_process_file.add_argument('--output-file', '-o', help='The location to output your data to', required=True)
    parser_process_file.add_argument('--output-format', '-o-form', help='The output format to be used, either csv or xlsx. Defaults to csv.', default="csv", const="csv", choices=["csv", "xlsx"], nargs="?")
    parser_process_file.add_argument('--coingecko-symbol-to-id-file', '-cgstoid', help='The location of your CoinGecko JSON config file that stores symbols to ID configurations for making API requests', required=True)
    parser_process_file.add_argument('--coinhall-symbol-to-id-file', '-chstoid', help='The location of your CoinHall JSON config file that stores symbols to ID configurations for making API requests', required=True)
    parser_process_file.add_argument('--symbol-column', '-sc', help='The column in the EXCEL that stores the symbol of interest', default="boughtCurrency")
    parser_process_file.add_argument('--value-column', '-vc', help='The column in the EXCEL that stores the symbol of interest', default="boughtQuantity")
    parser_process_file.add_argument('--date-column', '-dc', help='The column in the EXCEL that stores the symbol of interest', default="timeExecuted")
    parser_process_file.add_argument('--year-filter', '-yf', help='The year to gather data for', default=2022, type=int)
    parser_process_file.add_argument('--coingecko-cache', '-cg-c', help='The JSON file holding cached data from CoinGecko (used in subsequent runs)', default="./.coingecko_cache.json", type=str)
    parser_process_file.set_defaults(func=process_function)

    parser_import_symbol = subparser.add_parser("import-symbol")
    parser_import_symbol.set_defaults(func=import_symbol_function)
    parser_import_symbol.add_argument('symbol', help='The symbol to import. Options will be chosen that closely match this symbol and the CoinGecko IDs will be saved to your configuration file.')
    parser_import_symbol.add_argument("--type", choices=["coingecko"], help='The config file type, used to gather ID lists for giving choices.', required=True)
    parser_import_symbol.add_argument('--config-file', '-cf', help='The location of your JSON config file that stores symbols to ID configurations for making API requests', required=True)

    args = parser.parse_args()
    return args

def validate_process_args(args):
    if not exists(args.input_file):
        raise CaughtError(f"No file exists at {args.input_file} for the input data file, please enter a correct filepath")

    if exists(args.output_file):
        answer = input("Output file already exists, would you like to overwrite it? (Y/n) >> ")
        if answer != "Y":
            raise CaughtError("Exiting...")
    
    if not exists(args.coingecko_symbol_to_id_file):
        raise CaughtError(f"No file exists at {args.coingecko_symbol_to_id_file} for the CoinGecko symbol to id configuration file, please enter a correct filepath")
    
    if not exists(args.coinhall_symbol_to_id_file):
        raise CaughtError(f"No file exists at {args.coinhall_symbol_to_id_file} for the Coinhall symbol to id configuration file, please enter a correct filepath")

def validate_import_symbol_args(args):
    if not exists(args.config_file):
        raise CaughtError(f"No file exists at {args.config_file} for the config data file, please enter a correct filepath")

def get_coingecko_cache(coingecko_cache_file):
    if not exists(coingecko_cache_file):
        print(f"No CoinGecko price cache file found at {coingecko_cache_file}, creating...")
        with open(coingecko_cache_file, 'w') as fp:
            json.dump({}, fp)
        return {}
    else:
        print(f"Found CoinGecko cache file {coingecko_cache_file}, loading...")
        return json.load(open(coingecko_cache_file))

def write_coingecko_cache(data, coingecko_cache_file):
    with open(coingecko_cache_file, 'w') as fp:
        json.dump(data, fp)

def load_config_file(fname):
    return json.load(open(fname))

def save_config_file(fname, config_value):
    return json.dump(config_value, open(fname, 'w'), indent=4)

def parse_input_data(fname):
    print("Loading data from Excel file")
    wb = None
    try:
        wb = load_workbook(filename=fname)
    except InvalidFileException:
        raise CaughtError(f"Input file '{fname}' is not in the correct Excel format, please check your file and try again")

    ws = wb.active
    rows = ws.rows
    headers = [c.value for c in next(rows)]
    active_sheet_name = ws.title
    return active_sheet_name, headers, [row for row in iter_worksheet(ws)]

def iter_worksheet(worksheet):
    # It's necessary to get a reference to the generator, as 
    # `worksheet.rows` returns a new iterator on each access.
    rows = worksheet.rows

    # Get the header values as keys and move the iterator to the next item
    keys = [c.value for c in next(rows)]
    for row in rows:
        values = [c.value for c in row]
        yield dict(zip(keys, values))

def get_key_data_point_indexes(rows, dateColumn, yearFilter):
    print("Gathering key data points for classifications", KEY_CLASSIFICATIONS)
    return [index for index, row in enumerate(rows) if "classification" in row and row["classification"] in KEY_CLASSIFICATIONS and row[dateColumn].year == yearFilter and row["boughtCurrency"]]

def process_rows(rows, key_data_point_indexes, coingecko_symbols_to_id_configs, coinhall_symbols_to_id_configs, date_column, symbolColumn, valueColumn, simplified_rows_headers, coingecko_cache, coingecko_cache_file):

    print(f"Processing {len(key_data_point_indexes)} rows that matched the metric")

    symbols_to_dates_to_rows = {}

    #build data structure storing symbols to dates to row indexes
    #used for ease-of-access and processing
    for i in key_data_point_indexes:
        row = rows[i]
        boughtCurrency = row[symbolColumn]
        if boughtCurrency not in symbols_to_dates_to_rows:
            symbols_to_dates_to_rows[boughtCurrency] = {}

        date_key = row[date_column].strftime("%d-%m-%Y")

        if date_key not in symbols_to_dates_to_rows[boughtCurrency]:
            symbols_to_dates_to_rows[boughtCurrency][date_key] = [i]
        else:
            symbols_to_dates_to_rows[boughtCurrency][date_key].append(i)

    symbols_to_dates_to_costs = {}

    #gather the cached CoinGecko costs so we dont attempt to grab them again
    print("Checking CoinGecko cache file for cached data...")
    num_cached = 0
    for symbol in symbols_to_dates_to_rows:
        for date in symbols_to_dates_to_rows[symbol]:
            if symbol in coingecko_cache and date in coingecko_cache[symbol]:
                num_cached += 1
                if symbol not in symbols_to_dates_to_costs:
                    symbols_to_dates_to_costs[symbol] = {}
                symbols_to_dates_to_costs[symbol][date] = coingecko_cache[symbol][date]

    if num_cached > 0:
        print(f"Found {num_cached} entries, these will be skipped when reaching out to CoinGecko")
    else:
        print(f"No cached entries found, all data needs to be retrieved")

    #remove the cached entries
    for symbol in symbols_to_dates_to_costs:
        for date in symbols_to_dates_to_costs[symbol]:
            del symbols_to_dates_to_rows[symbol][date]
        # if len(symbols_to_dates_to_rows[symbol].keys()) == 0:
        #     del symbols_to_dates_to_rows[symbol]
    
    #loop through chunks of data and send to API for processing
    print("Reaching out to CoinGecko for symbol cost data, this may take a while...")

    coingecko_requests_made = 0
    for symbol in symbols_to_dates_to_rows:

        if symbol not in symbols_to_dates_to_costs:
                symbols_to_dates_to_costs[symbol] = {}

        request_url = get_coingecko_request_url(symbol, coingecko_symbols_to_id_configs)

        for date in symbols_to_dates_to_rows[symbol]:
            if symbol in coingecko_cache and date in coingecko_cache[symbol]:
                symbols_to_dates_to_costs[symbol][date] = coingecko_cache[symbol][date]
                continue
            else:
                symbols_to_dates_to_costs[symbol][date] = None
                if request_url:
                    parameterized_url = add_coingecko_request_params(request_url, date)
                    resp = requests.get(parameterized_url)

                    try:
                        resp.raise_for_status()
                    except Exception as err:
                        if resp.status_code == 429:
                            print("A CoinGecko API throttle error occurred, self-throttling and trying again")
                            time.sleep(COIN_GECKO_THROTTLE_TIME)
                            resp = requests.get(parameterized_url)
                            coingecko_requests_made = 1

                            #give up after 1 retry
                            try:
                                resp.raise_for_status()
                            except Exception as err:
                                write_coingecko_cache(coingecko_cache, coingecko_cache_file)
                                print("CoinGecko API call failed", err)
                                raise CaughtError(f"CoinGecko API call failed, subsequent runs will use the cache file at {coingecko_cache_file} to start from this checkpoint")
                        else:
                            write_coingecko_cache(coingecko_cache, coingecko_cache_file)
                            print("CoinGecko API call failed", err)
                            raise CaughtError(f"CoinGecko API call failed, subsequent runs will use the cache file at {coingecko_cache_file} to start from this checkpoint")
                        
                    #safely get a None value if the json structure is not found in response
                    #this was found during testing of some symbols where they store data but not the USD cost on that date
                    symbol_date_cost = resp.json().get("market_data", {}).get("current_price", {}).get("usd")
                    symbols_to_dates_to_costs[symbol][date] = symbol_date_cost

                    if symbol in coingecko_cache:
                        coingecko_cache[symbol][date] = symbols_to_dates_to_costs[symbol][date]
                    else:
                        coingecko_cache[symbol] = {}
                        coingecko_cache[symbol][date] = symbols_to_dates_to_costs[symbol][date]

                    write_coingecko_cache(coingecko_cache, coingecko_cache_file)

                    coingecko_requests_made += 1
                    #self-throttling to stay under CoinGecko throttle limits
                    if coingecko_requests_made == COIN_GECKO_REQUEST_CHUNKS:
                        coingecko_requests_made = 0
                        time.sleep(COIN_GECKO_THROTTLE_TIME)
                    else:
                        time.sleep(1)
                else:
                    print("Your CoinGecko symbol config list does not support symbol", symbol)

    write_coingecko_cache(coingecko_cache, coingecko_cache_file)
    print("Finished reaching out to CoinGecko")

    missing_coingecko_coverage = {}

    for i in key_data_point_indexes:
        row = rows[i]

        boughtCurrency = row[symbolColumn]
        date_key = row[date_column].strftime("%d-%m-%Y")

        try:
            row["usdValue"] = symbols_to_dates_to_costs[boughtCurrency][date_key] * row[valueColumn]
        except:
            missing_coingecko_coverage[i] = {"symbol": boughtCurrency, "date": row[date_column]}

        rows[i] = row

    missing_coinhall_coverage = {}
    if missing_coingecko_coverage:
        print("CoinGecko does not provide coverage for the following Symbol + Date combinations:")
        for index in missing_coingecko_coverage:
            print(f"\tSymbol: {missing_coingecko_coverage[index]['symbol']}\t\tDate: {missing_coingecko_coverage[index]['date']}\t\t(Excel Row {index + 2})")

        print("Attempting Coinhall fallback API")

        index_to_costs = {}
        print("Reaching out to Coinhall for symbol cost data, this may take a while...")
        for index in missing_coingecko_coverage:
            symbol = missing_coingecko_coverage[index]['symbol']
            date = missing_coingecko_coverage[index]['date']
            
            request_url = get_coinhall_request_url(symbol, coinhall_symbols_to_id_configs)

            if request_url:
                parameterized_url = add_coinhall_request_params(request_url, date, coinhall_symbols_to_id_configs[symbol]["id"])
                resp = make_coinhall_api_request(parameterized_url, "Error reaching out to CoinGecko, please try again later:", COINHALL_THROTTLE_TIME)
                data = resp.json()
                if data and len(data) > 0:
                    index_to_costs[index] = data[0]
                else:
                    missing_coinhall_coverage[index] = {"symbol": symbol, "date": date}
                time.sleep(COINHALL_THROTTLE_TIME)
            else:
                print("Your Coinhall symbol config list does not support symbol", symbol)
                missing_coinhall_coverage[index] = {"symbol": symbol, "date": date}
        
        print("Finished reaching out to Coinhall")
        for index in index_to_costs:
            row = rows[index]
            row["usdValue"] = index_to_costs[index]["high"] * row[valueColumn]
            rows[index] = row

        if missing_coinhall_coverage:
            print("CoinHall does not provide coverage for the following Symbol + Date combinations:")
            for index in missing_coinhall_coverage:
                row = rows[index]
                row["usdValue"] = 0
                rows[index] = row
                print(f"\tSymbol: {missing_coinhall_coverage[index]['symbol']}\t\tDate: {missing_coinhall_coverage[index]['date']}\t\t(Excel Row {index + 2})")
        else:
            print("CoinHall provided coverage for all CoinGecko missing symbol/date combinations")
    else:
        print("CoinGecko provided coverage for all symbol/date combinations")
    

    simplified_rows = []

    for index in key_data_point_indexes:
        row = rows[index]
        simplified_row = {header: row[header] for header in simplified_rows_headers}
        if index in missing_coinhall_coverage:
            simplified_row["comment"] = "Not covered, fixme"
        else:
            simplified_row["comment"] = ""
        simplified_rows.append(simplified_row)



    return rows, simplified_rows

#creates chunks of symbols and dates to send for API processing
def chunk_symbols_and_dates(symbols_to_dates, len_chunks):

    values = {}
    length = 0
    for key in symbols_to_dates:
        for date in symbols_to_dates[key]:
            if key in values:
                values[key].append(date)
                length+=1
            else:
                values[key] = [date]
                length+=1
            if length == len_chunks:
                yield values
                values = {}
                length = 0

    #last yield returns last chunk
    yield values

def get_coingecko_request_url(symbol, coingecko_symbols_to_id_configs):
    coingecko_request_id = None
    coingecko_request_url = None

    if symbol in coingecko_symbols_to_id_configs:
        coingecko_request_id = coingecko_symbols_to_id_configs[symbol]["id"]
        coingecko_request_url =  COIN_GECKO_API_URL + COIN_GECKO_COINS_ENDPOINT + COIN_GECKO_HISTORICAL_ENDPOINT.replace("<coin-id>", coingecko_request_id)

    return coingecko_request_url

def add_coingecko_request_params(request_url, date):
    return request_url + f"?date={date}"

def get_coinhall_request_url(symbol, coinhall_symbols_to_id_configs):
    coinhall_request_url = None

    if symbol in coinhall_symbols_to_id_configs:
        coinhall_request_url =  COINHALL_API_URL + COINHALL_CHART_ENDPOINT

    return coinhall_request_url

def add_coinhall_request_params(request_url, date, symbol_id):
    from_date = date.replace(hour=0, minute=0, second=0, microsecond=0)
    to_date = from_date.replace(hour=11, minute=59, second=59)

    from_unix = str(int(time.mktime(from_date.timetuple())))
    to_unix = str(int(time.mktime(to_date.timetuple())))

    return request_url + COINHALL_HISTORICAL_PARAMS.replace("<from-time>", from_unix).replace("<to-time>", to_unix).replace("<coin-id>", symbol_id)

def make_coinhall_api_request(request_url, error_string, throttle_time):
    resp = requests.get(request_url)
    try:
        resp.raise_for_status()
    except requests.HTTPError as err:
        #if throttled, wait for the throttle to end and retry
        if resp.status_code == 429:
            time.sleep(throttle_time)
            resp = requests.get(request_url)

            #give up after 1 retry
            try:
                resp.raise_for_status()
            except Exception as err:
                print(error_string, err)
                sys.exit(1)
        else:
            print(error_string, err)
            sys.exit(1)
    except Exception as err:
        print(error_string, err)
        sys.exit(1)

def output_rows(title, headers, rows, fname, format, sum_header=None):
    if format == "xlsx":
        print("Creating new Excel file", fname)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = title

        ws1.append(headers)

        for row in rows:
            data = []
            for header in headers:
                if header in row:
                    data.append(row[header])
                else:
                    data.append(None)
            ws1.append(data)

        if sum_header and sum_header in headers:
            index = headers.index(sum_header) + 1
            column = get_column_letter(index)

            first_row = column + "2"
            last_row = column + str(len(ws1[column]))

            sum_formula = f"= SUM({first_row}:{last_row})"

            sum_formula_cell = column + str(len(ws1[column]) + 1)
            ws1[sum_formula_cell] = sum_formula

        wb.save(fname)

    elif format == "csv":
        print(f"Creating new CSV file {fname}")
        with open(fname, 'w') as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerow(headers)
            for row in rows:
                data = []
                for header in headers:
                    if header in row:
                        data.append(row[header])
                    else:
                        data.append("")
                csvwriter.writerow(data)

def process(args):

    validate_process_args(args)

    coingecko_cache = get_coingecko_cache(args.coingecko_cache)

    #Curated list of CoinGecko Symbols to IDs in a JSON file
    #These will correspond to the symbols found in your CSV column -> CoinGecko's data structure for corresponding symbol
    #e.g.
    # {
    #     "DAI": {
    #         "id": "dai",
    #         "symbol": "dai",
    #         "name": "Dai"
    #     },
    #     "ADA": {
    #         "id": "cardano",
    #         "symbol": "ada",
    #         "name": "Cardano"
    #     },
    #     ...
    # }
    coingecko_symbols_to_id_configs = load_config_file(args.coingecko_symbol_to_id_file)

    #Curated list of Coinhall Symbols to IDs in a JSON file
    #Coinhall provides coverage for various Terra Symbols by pairAddress
    #These will correspond to the symbols found in your CSV column -> Coinhalls's data structure for corresponding symbol
    #e.g.
    # {
    #      "MINE": {
    #         "id": "terra178jydtjvj4gw8earkgnqc80c3hrmqj4kw2welz",
    #         "symbol": "mine",
    #         "name": "Pylon Protocol"
    #     },
    #     "PSI": {
    #         "id": "terra163pkeeuwxzr0yhndf8xd2jprm9hrtk59xf7nqf",
    #         "symbol": "psi",
    #         "name": "Nexus Protocol"
    #     },
    #     "LOOP": {
    #         "id": "terra106a00unep7pvwvcck4wylt4fffjhgkf9a0u6eu",
    #         "symbol": "loop",
    #         "name": "LOOP Finance"
    #     },
    #     ...
    # }
    coinhall_symbols_to_id_configs = load_config_file(args.coinhall_symbol_to_id_file)

    title, headers, rows = parse_input_data(args.input_file)
    key_data_point_indexes = get_key_data_point_indexes(rows, args.date_column, args.year_filter)
    simplified_rows_headers = [args.date_column, args.symbol_column, args.value_column, "usdValue"]
    rows, simplified_rows = process_rows(rows, key_data_point_indexes, coingecko_symbols_to_id_configs, coinhall_symbols_to_id_configs, args.date_column, args.symbol_column, args.value_column, simplified_rows_headers, coingecko_cache, args.coingecko_cache)
    headers.append("usdValue")
    simplified_rows_headers.append("comment")
    output_rows(title, headers, rows, args.output_file, args.output_format)
    output_rows(title, simplified_rows_headers, simplified_rows, args.output_file + f"-simplified.{args.output_format}", args.output_format, sum_header="usdValue")

def import_symbol_coingecko_worker(symbol, config_file):
    print(f"Searching CoinGecko for symbol {symbol}")

    #TODO: Cache this as well for later use? Would want to stat the config file and reload the list if its pretty old
    resp = requests.get(COIN_GECKO_API_URL + COIN_GECKO_COINS_ENDPOINT + COIN_GECKO_COINS_LIST_ENDPOINT)
    try:
        resp.raise_for_status()
    except requests.HTTPError as err:
        print(err)
        raise CaughtError("Error reaching out to CoinGecko API List endpoint, please try again in a minute or two.")
    except Exception as err:
        raise err

    data = resp.json()

    possible_options = []
    #basic O(n*m) string search, super slow
    for symbol_config in data:
        if symbol in symbol_config["symbol"] or symbol.lower() in symbol_config["symbol"]:
            possible_options.append(symbol_config)

    if len(possible_options) == 0:
        print(f"No options found that match symbol {symbol}")
        print("If this sounds like an error, you can try manually looking through the JSON file provided by CoinGecko at:")
        print(COIN_GECKO_API_URL + COIN_GECKO_COINS_ENDPOINT + COIN_GECKO_COINS_LIST_ENDPOINT)
        return
    
    print(f"Found {len(possible_options)} possible options for symbol {symbol}")

    print("Ranking them by least likely to most likely, this may take a second...")

    #Rank possible options from least likely to most likely
    possible_option_ranks = []
    for possible_option in possible_options:
        dist = levenshtein_dist_dp(symbol.lower(), possible_option["symbol"])
        possible_option_ranks.append((possible_option, dist))
    
    #sort by rank
    possible_option_ranks = sorted(possible_option_ranks, key=lambda possible_option: possible_option[1])

    #build user input information string
    dump_string = "These are the possible options (press 'q' to continue):\n\n"
    for idx, option in enumerate(possible_option_ranks):
        possible_option = option[0]
        dump_string += f"{idx + 1}: {possible_option['name']}\n\tSymbol: {possible_option['symbol']}\n\tID: {possible_option['id']}\n"

    while(True):
        #paging output like the Linux less command
        pydoc.pager(dump_string)

        answer = input("Which option would you like to choose? (input the number or type 'ls' to show again) >> ")

        if answer == "ls":
            continue
        else:
            while True:
                try:
                    val = int(answer)
                    if val - 1 < 0:
                        raise IndexError
                    chosen = possible_option_ranks[val - 1][0]
                    break
                except ValueError:
                    answer = input("Please enter an integer value >> ")
                    continue
                except IndexError:
                    answer = input(f"{answer} is not a valid option in the list, please choose again >> ")

        print("You have chosen the following option:")
        print(f"{val}: {chosen['name']}\n\tSymbol: {chosen['symbol']}\n\tID: {chosen['id']}")
        done = False
        while True:
            answer = input("Is this correct? (y/n) >> ")
            if answer == "y":
                done = True
                break
            elif answer == "n":
                break
            else:
                print("Please type y/n")
        if done:
            break
    
    print(f"Saving the config {json.dumps(chosen)} into your symbol to IDs config file '{config_file}' at key {symbol}")

    current_config = load_config_file(config_file)

    #ask before overwrite
    if symbol in current_config:
        print(f"Your config file already contains an entry for symbol {symbol} as {current_config[symbol]}")
        while(True):
            answer = input("Would you like to overwrite this entry? (y/n) >> ")
            if answer == "y" or answer == "n":
                break
        if answer == "n":
            print("Exiting...")
            return
        else:
            print("Overwriting...")

    current_config[symbol] = chosen

    save_config_file(config_file, current_config)
    print("Done!")

def import_symbol(args):
    validate_import_symbol_args(args)

    symbol = args.symbol
    api_type = args.type

    config_file = args.config_file

    #TODO: Add coinhall?
    if api_type == "coingecko":
        import_symbol_coingecko_worker(symbol, config_file)
            
#Pulled from https://github.com/pharr117/levenshtein_dist 
#A general string comparison algorithm: given a source string and a target string, calculates the distance between them
#Returns a rank, the lower the rank the closer the words match.
def levenshtein_dist_dp(source, target):

    len_source = len(source)
    len_target = len(target)
    matrix = [[None for j in range(len_target+1)] for i in range(len_source+1)]

    for i in range(len_source + 1):
        for j in range(len_target+1):
            if i == 0:
                matrix[i][j] = j
            elif j == 0:
                matrix[i][j] = i
            else:
                if source[i-1] == target[j-1]:
                    matrix[i][j] = matrix[i-1][j-1]
                else:
                    matrix[i][j] = 1 + min(matrix[i][j-1],
                                           matrix[i-1][j],
                                           matrix[i-1][j-1])

    return matrix[len_source][len_target]

def main():
    args = parse_args(process, import_symbol)
    args.func(args)

if __name__ == "__main__":
    try:
        main()
    except CaughtError as err:
        print(err)
    except Exception as err:
        print("An unanticipated error occured, please contact the developer to assist (provide the following stack trace)")
        traceback.print_exc()
        print(err)